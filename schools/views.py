import re
import string
from pathlib import Path

from django.conf import settings
from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from users.models import User, UserRole

from .models import Classroom, ClassroomUser, School
from .permissions import (IsSchoolAdminOrSuperAdmin, IsSuperAdmin,
                          IsTeacherOrAbove)
from .serializers import (BulkClassroomUserSerializer,
                          ClassroomDetailSerializer, ClassroomSerializer,
                          ClassroomUserSerializer, SchoolSerializer)


class SchoolViewSet(viewsets.ModelViewSet):
    queryset = School.objects.all()
    serializer_class = SchoolSerializer
    permission_classes = [IsSuperAdmin]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['name', 'city', 'country']
    ordering_fields = ['name', 'city']
    ordering = ['name']

    @action(detail=True, methods=['get'], url_path='credentials-files')
    def credentials_files(self, request, pk=None):
        """
        List all generated credentials Excel files for this school (students and teachers).
        """
        school = self.get_object()
        base_dir = Path(settings.MEDIA_ROOT) / "import-credentials"

        def collect_files(subdir: str):
            items = []
            target = base_dir / subdir
            if not target.exists():
                return items
            for path in sorted(target.glob(f"*_{school.id}_*.xlsx")):
                stat = path.stat()
                created = timezone.datetime.fromtimestamp(stat.st_mtime, tz=timezone.get_current_timezone())
                rel = path.relative_to(settings.MEDIA_ROOT)
                absolute_url = request.build_absolute_uri(f"{settings.MEDIA_URL}{rel.as_posix()}")
                items.append(
                    {
                        "filename": path.name,
                        "url": absolute_url,
                        "created_at": created.isoformat(),
                        "size": stat.st_size,
                    }
                )
            return items

        return Response(
            {
                "students": collect_files("students"),
                "teachers": collect_files("teachers"),
            }
        )

    @action(detail=True, methods=['post'], url_path='import-teachers-excel')
    def import_teachers_excel(self, request, pk=None):
        """
        Import teachers from Excel file.
        Supported column names (Russian or English):
          - "ФИО Учителя" / "ФИ Учителя" / "teacher_full_name" — full name (split into last + first)
          - "first_name" / "Имя"
          - "last_name" / "Фамилия"
          - "Эл.почта" / "email" / "Email"
          - "phone_number" / "Телефон"
        Query/body: preview=1 — parse-only, no DB writes.
        """
        school = self.get_object()
        is_preview = request.data.get('preview') in ('1', 'true', True)

        if 'file' not in request.FILES:
            return Response({'error': 'Excel file is required'}, status=status.HTTP_400_BAD_REQUEST)

        excel_file = request.FILES['file']
        default_password = request.data.get('default_password', None)

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': 'File must be an Excel file (.xlsx or .xls)'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            import openpyxl
        except ImportError:
            return Response({'error': 'openpyxl is required'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            workbook = openpyxl.load_workbook(excel_file, read_only=True)

            def _norm_teacher_header(raw: str) -> str:
                # Normalize: strip, collapse all whitespace variants (incl. \xa0) to plain space
                key = ' '.join(str(raw).strip().replace('\xa0', ' ').split()).lower()
                if key in {'first_name', 'имя'}:
                    return 'first_name'
                if key in {'last_name', 'фамилия'}:
                    return 'last_name'
                if 'фио учителя' in key or 'фи учителя' in key or 'teacher_full_name' in key:
                    return 'teacher_full_name'
                if key in {'email', 'e-mail'} or 'почта' in key:
                    return 'email'
                if key in {'phone_number', 'phone'} or 'телефон' in key:
                    return 'phone_number'
                return key

            def _is_teacher_header_row(normalized_cells: list[str]) -> bool:
                keywords = {'first_name', 'last_name', 'email', 'teacher_full_name', 'phone_number'}
                return bool(keywords & set(normalized_cells))

            def split_full_name(full_name: str) -> tuple[str | None, str | None]:
                parts = [p for p in str(full_name).strip().split() if p]
                if not parts:
                    return None, None
                if len(parts) == 1:
                    return parts[0], ''
                return parts[0], parts[1]

            def _find_headers(worksheet):
                for row_idx, row in enumerate(
                    worksheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1
                ):
                    if not row:
                        continue
                    normed = [_norm_teacher_header(c) for c in row if c]
                    if not normed:
                        continue
                    if _is_teacher_header_row(normed):
                        hdrs = {}
                        for col_idx, cell in enumerate(row, start=1):
                            if cell:
                                hdrs[_norm_teacher_header(cell)] = col_idx
                        return hdrs, row_idx
                return {}, None

            def _extract_row(row, headers):
                def _val(key):
                    idx = headers.get(key)
                    if idx and idx <= len(row) and row[idx - 1]:
                        return str(row[idx - 1]).replace('\xa0', ' ').strip()
                    return None

                first_name = last_name = None
                if 'teacher_full_name' in headers:
                    raw = _val('teacher_full_name')
                    if raw:
                        last_name, first_name = split_full_name(raw)
                else:
                    first_name = _val('first_name')
                    last_name = _val('last_name')

                email = _val('email')
                phone = _val('phone_number')
                return first_name, last_name, email, phone

            # ── PREVIEW ────────────────────────────────────────────────────────
            if is_preview:
                preview_rows: list[dict] = []
                teachers_new = 0
                teachers_existing = 0
                preview_errors: list[dict] = []

                for worksheet in workbook.worksheets:
                    headers, header_row = _find_headers(worksheet)
                    if not headers or header_row is None:
                        continue
                    if 'teacher_full_name' not in headers and ('first_name' not in headers or 'last_name' not in headers):
                        continue

                    for row_idx, row in enumerate(
                        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                        start=header_row + 1,
                    ):
                        if not any(cell for cell in row):
                            continue

                        first_name, last_name, email, _ = _extract_row(row, headers)

                        if not last_name:
                            continue

                        existing = User.objects.filter(
                            first_name=first_name,
                            last_name=last_name,
                            school=school,
                            role=UserRole.TEACHER,
                        ).first()

                        if existing:
                            teachers_existing += 1
                            teacher_status = 'existing'
                        else:
                            teachers_new += 1
                            teacher_status = 'new'

                        preview_rows.append({
                            'row': row_idx,
                            'first_name': first_name,
                            'last_name': last_name,
                            'email': email,
                            'teacher_status': teacher_status,
                        })

                return Response({
                    'preview': True,
                    'summary': {
                        'teachers_new': teachers_new,
                        'teachers_existing': teachers_existing,
                        'rows_count': len(preview_rows),
                        'errors_count': len(preview_errors),
                    },
                    'rows': preview_rows[:100],
                    'errors': preview_errors[:50],
                }, status=status.HTTP_200_OK)

            # ── REAL IMPORT ────────────────────────────────────────────────────
            if not (default_password and str(default_password).strip()):
                default_password = 'qwerty123'

            results = {
                'created_teachers': 0,
                'skipped_existing': 0,
                'errors': [],
                'default_password': default_password,
                'new_teachers_credentials': [],
            }
            credentials_file_path = None

            with transaction.atomic():
                for worksheet in workbook.worksheets:
                    headers, header_row = _find_headers(worksheet)
                    if not headers or header_row is None:
                        continue
                    if 'teacher_full_name' not in headers and ('first_name' not in headers or 'last_name' not in headers):
                        continue

                    for row_idx, row in enumerate(
                        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                        start=header_row + 1,
                    ):
                        if not any(cell for cell in row):
                            continue

                        first_name, last_name, email, phone = _extract_row(row, headers)

                        if not last_name:
                            continue

                        try:
                            with transaction.atomic():
                                existing = User.objects.filter(
                                    first_name=first_name,
                                    last_name=last_name,
                                    school=school,
                                    role=UserRole.TEACHER,
                                ).first()

                                if existing:
                                    results['skipped_existing'] += 1
                                    continue

                                username = self._generate_username(first_name or '', last_name, school)
                                user = User.objects.create_user(
                                    username=username,
                                    email=email or None,
                                    password=default_password,
                                    role=UserRole.TEACHER,
                                    first_name=first_name or '',
                                    last_name=last_name,
                                    phone_number=phone or None,
                                    school=school,
                                    is_active=True,
                                    must_change_password=True,
                                )
                                results['created_teachers'] += 1
                                results['new_teachers_credentials'].append({
                                    'first_name': first_name or '',
                                    'last_name': last_name,
                                    'email': email,
                                    'username': username,
                                    'password': default_password,
                                })

                        except Exception as e:
                            results['errors'].append({'row': row_idx, 'error': str(e)})

            # Generate credentials Excel
            if results['new_teachers_credentials']:
                try:
                    import openpyxl as _xl
                    from openpyxl.utils import get_column_letter

                    wb = _xl.Workbook()
                    ws = wb.active
                    ws.title = 'Teachers credentials'
                    hdr = ['Фамилия', 'Имя', 'Email', 'Логин', 'Пароль']
                    ws.append(hdr)
                    for cred in results['new_teachers_credentials']:
                        ws.append([
                            cred['last_name'],
                            cred['first_name'],
                            cred['email'] or '',
                            cred['username'],
                            cred['password'],
                        ])
                    for col_idx, title in enumerate(hdr, start=1):
                        max_len = len(title)
                        for r in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
                            if r[0]:
                                max_len = max(max_len, len(str(r[0])))
                        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

                    from pathlib import Path
                    target_dir = Path(settings.MEDIA_ROOT) / 'import-credentials' / 'teachers'
                    target_dir.mkdir(parents=True, exist_ok=True)
                    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                    file_name = f'teachers_credentials_school_{school.id}_{timestamp}.xlsx'
                    credentials_file_path = target_dir / file_name
                    wb.save(credentials_file_path)
                except Exception as e:
                    results['errors'].append({'row': None, 'error': f'Credentials Excel failed: {e}'})

            response_data = {
                'success': True,
                'message': 'Import completed',
                'summary': {
                    'total_teachers': results['created_teachers'],
                    'skipped_existing': results['skipped_existing'],
                    'errors_count': len(results['errors']),
                },
                'default_password': results['default_password'],
                'errors': results['errors'][:50],
            }
            if credentials_file_path is not None:
                from pathlib import Path as _Path
                rel = _Path(credentials_file_path).relative_to(settings.MEDIA_ROOT)
                response_data['credentials_file'] = {
                    'path': str(rel),
                    'url': f"{settings.MEDIA_URL}{rel.as_posix()}",
                }

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': f'Error processing file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='import-students-excel')
    def import_students_excel(self, request, pk=None):
        """
        Import students from Excel file.
        Expected columns: class_name, first_name, last_name, email (optional), phone_number (optional),
        parent_username (optional, per row). If parent_username column is missing/empty, form field is used.
        Optional form field: parent_username — default for rows without parent_username in Excel.
        Query/body: preview=1 — only parse and return counts (no DB writes).
        """
        school = self.get_object()
        is_preview = request.data.get('preview') in ('1', 'true', True)

        # Check if file is provided
        if 'file' not in request.FILES:
            return Response(
                {'error': 'Excel file is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        excel_file = request.FILES['file']
        default_password = request.data.get('default_password', None)
        default_parent_username = (request.data.get('parent_username') or '').strip() or None

        # Validate file extension
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'File must be an Excel file (.xlsx or .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            import openpyxl
        except ImportError:
            return Response(
                {'error': 'openpyxl library is required. Install it with: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        try:
            # Load workbook
            workbook = openpyxl.load_workbook(excel_file, read_only=True)

            def _normalize_header_name(raw: str) -> str:
                key = ' '.join(str(raw).strip().replace('\xa0', ' ').split()).lower()
                # English / Russian keys
                if key in {"class_name", "class", "класс"}:
                    return "class_name"
                if key in {"first_name", "имя"}:
                    return "first_name"
                if key in {"last_name", "фамилия"}:
                    return "last_name"
                # Full name of student
                if "фи ученика" in key or "фио ученика" in key:
                    return "student_full_name"
                # Parent username (old format)
                if key == "parent_username":
                    return "parent_username"
                # Parent full name (new format)
                if "фио родителя" in key or "фи родителя" in key:
                    return "parent_full_name"
                # Email
                if key in {"email", "e-mail"} or "почта" in key:
                    return "email"
                # Phone
                if key in {"phone_number", "phone"} or "телефон" in key:
                    return "phone_number"
                return key

            # Helper: split full name like "Фамилия Имя Отчество" -> (Фамилия, Имя)
            def split_full_name(full_name: str) -> tuple[str | None, str | None]:
                if not full_name:
                    return None, None
                parts = [p for p in str(full_name).strip().split() if p]
                if not parts:
                    return None, None
                if len(parts) == 1:
                    return parts[0], ''
                # Ignore отчество и остальные части
                last_name = parts[0]
                first_name = parts[1]
                return last_name, first_name

            # Helper: per-row parent username from legacy column
            def row_parent_username(row, row_len, headers, default_username):
                if (
                    "parent_username" in headers
                    and headers["parent_username"] <= row_len
                    and row[headers["parent_username"] - 1]
                ):
                    return (
                        str(row[headers["parent_username"] - 1]).strip()
                        or default_username
                    )
                return default_username

            # Preview mode: aggregate across all sheets
            if is_preview:
                preview_rows: list[dict] = []
                students_new = 0
                students_existing = 0
                parents_new_count = 0
                parents_existing_count = 0
                preview_errors: list[dict] = []

                for worksheet in workbook.worksheets:
                    headers: dict[str, int] = {}
                    header_row: int | None = None

                    # Find header row on this sheet
                    for row_idx, row in enumerate(
                        worksheet.iter_rows(min_row=1, max_row=10, values_only=True),
                        start=1,
                    ):
                        if not row:
                            continue
                        normalized_cells = [str(cell).strip().lower() for cell in row if cell]
                        if not normalized_cells:
                            continue
                        if any(
                            c in ["class_name", "first_name", "last_name", "класс"]
                            or "фи ученика" in c
                            or "фио ученика" in c
                            or "почта" in c
                            or "фио родителя" in c
                            or "фи родителя" in c
                            for c in normalized_cells
                        ):
                            for col_idx, cell_value in enumerate(row, start=1):
                                if cell_value:
                                    norm_key = _normalize_header_name(cell_value)
                                    headers[norm_key] = col_idx
                            header_row = row_idx
                            break

                    if not headers or header_row is None:
                        # This sheet does not contain a recognizable header, skip it
                        continue

                    if "class_name" not in headers:
                        # Not a student import sheet
                        continue

                    has_separate_names = "first_name" in headers and "last_name" in headers
                    has_full_name = "student_full_name" in headers
                    if not has_separate_names and not has_full_name:
                        # This sheet does not have required name columns, skip it
                        continue

                    # Per-sheet set for parent logical keys
                    parent_usernames_seen = set()

                    for row_idx, row in enumerate(
                        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                        start=header_row + 1,
                    ):
                        if not any(cell for cell in row):
                            continue

                        class_name = (
                            str(row[headers["class_name"] - 1]).strip()
                            if headers["class_name"] <= len(row)
                            and row[headers["class_name"] - 1]
                            else None
                        )

                        # Extract student name depending on format
                        first_name = None
                        last_name = None
                        if "student_full_name" in headers and headers["student_full_name"] <= len(row):
                            full_name_raw = row[headers["student_full_name"] - 1]
                            if full_name_raw:
                                last_name, first_name = split_full_name(full_name_raw)
                        else:
                            first_name = (
                                str(row[headers["first_name"] - 1]).strip()
                                if headers.get("first_name", 0) <= len(row)
                                and row[headers["first_name"] - 1]
                                else None
                            )
                            last_name = (
                                str(row[headers["last_name"] - 1]).strip()
                                if headers.get("last_name", 0) <= len(row)
                                and row[headers["last_name"] - 1]
                                else None
                            )

                        email = (
                            str(row[headers["email"] - 1]).strip()
                            if "email" in headers
                            and headers["email"] <= len(row)
                            and row[headers["email"] - 1]
                            else None
                        )

                        # Считаем только строки, где есть ФИ ученика и класс
                        if not class_name or not last_name:
                            continue

                        grade, letter = self._parse_class_name(class_name)
                        if not grade or not letter:
                            preview_errors.append(
                                {
                                    "row": row_idx,
                                    "error": f"Invalid class_name: {class_name}",
                                }
                            )
                            continue

                        # Determine if student already exists: prefer email, fallback to name + school
                        if email:
                            existing_student = User.objects.filter(email=email).first()
                        else:
                            existing_student = User.objects.filter(
                                first_name=first_name,
                                last_name=last_name,
                                school=school,
                                role=UserRole.STUDENT,
                            ).first()
                        student_exists = existing_student is not None
                        if student_exists:
                            students_existing += 1
                        else:
                            students_new += 1

                        # Parent detection: either by username (old format) or by full name (new format)
                        p_username = None
                        parent_status = None

                        if "parent_username" in headers:
                            p_username = row_parent_username(
                                row, len(row), headers, default_parent_username
                            )
                            if p_username:
                                if p_username not in parent_usernames_seen:
                                    parent_usernames_seen.add(p_username)
                                    par = User.objects.filter(
                                        username=p_username, role=UserRole.PARENT
                                    ).first()
                                    if par:
                                        parents_existing_count += 1
                                        parent_status = "existing"
                                    elif User.objects.filter(username=p_username).exists():
                                        parent_status = "exists_not_parent"
                                    else:
                                        parents_new_count += 1
                                        parent_status = "new"
                                else:
                                    par = User.objects.filter(
                                        username=p_username, role=UserRole.PARENT
                                    ).first()
                                    parent_status = "existing" if par else "new"
                        elif (
                            "parent_full_name" in headers
                            and headers["parent_full_name"] <= len(row)
                        ):
                            p_full = row[headers["parent_full_name"] - 1]
                            if p_full:
                                pl, pf = split_full_name(p_full)
                                if pl:
                                    logical_key = f"{pl} {pf}".strip()
                                    p_username = logical_key
                                    if logical_key not in parent_usernames_seen:
                                        parent_usernames_seen.add(logical_key)
                                        par = User.objects.filter(
                                            first_name=pf,
                                            last_name=pl,
                                            role=UserRole.PARENT,
                                        ).first()
                                        if par:
                                            parents_existing_count += 1
                                            parent_status = "existing"
                                        else:
                                            parents_new_count += 1
                                            parent_status = "new"
                                    else:
                                        par = User.objects.filter(
                                            first_name=pf,
                                            last_name=pl,
                                            role=UserRole.PARENT,
                                        ).first()
                                        parent_status = "existing" if par else "new"

                        preview_rows.append(
                            {
                                "row": row_idx,
                                "class_name": class_name,
                                "first_name": first_name,
                                "last_name": last_name,
                                "email": email,
                                "student_status": "existing"
                                if student_exists
                                else "new",
                                "parent_username": p_username or None,
                                "parent_status": parent_status,
                            }
                        )

                return Response(
                    {
                        "preview": True,
                        "summary": {
                            "students_new": students_new,
                            "students_existing": students_existing,
                            "parents_new": parents_new_count,
                            "parents_existing": parents_existing_count,
                            "rows_count": len(preview_rows),
                            "errors_count": len(preview_errors),
                        },
                        "rows": preview_rows[:100],
                        "errors": preview_errors[:50],
                    },
                    status=status.HTTP_200_OK,
                )

            # Default password for all created students and new parents
            if not (default_password and str(default_password).strip()):
                default_password = 'qwerty123'

            # Global accumulators for real import across all sheets
            results = {
                "created_classrooms": {},
                "created_students": {},
                "errors": [],
                "default_password": default_password,
                "new_students_credentials": [],
            }

            # Cache: logical parent key -> parent User (for per-row parent data)
            parent_cache = {}
            created_parent_passwords = {}
            credentials_file_path = None

            def get_or_create_parent(
                p_key, first_name: str | None = None, last_name: str | None = None
            ):
                """
                p_key:
                    - old формат: username строки
                    - новый формат: логический ключ по ФИ родителя ("Фамилия Имя")
                first_name / last_name используются только в новом формате для установки имени.
                """
                if not p_key:
                    return None, None
                if p_key in parent_cache:
                    return parent_cache[p_key], created_parent_passwords.get(p_key)

                # Новый формат: поиск по имени/фамилии родителя
                if first_name or last_name:
                    qs = User.objects.filter(role=UserRole.PARENT)
                    if first_name:
                        qs = qs.filter(first_name=first_name)
                    if last_name:
                        qs = qs.filter(last_name=last_name)
                    par = qs.first()
                    if par:
                        parent_cache[p_key] = par
                        return par, None
                else:
                    # Старый формат: поиск по username
                    par = User.objects.filter(
                        username=p_key, role=UserRole.PARENT
                    ).first()
                    if par:
                        parent_cache[p_key] = par
                        return par, None

                if not (first_name or last_name):
                    existing = User.objects.filter(username=p_key).first()
                    if existing:
                        return None, "not_parent"  # signal error

                pwd = default_password
                # Create parent user without synthetic email if none is provided
                if first_name or last_name:
                    username = self._generate_username(
                        first_name or "", last_name or "", school
                    )
                    par = User.objects.create_user(
                        username=username,
                        email=None,
                        password=pwd,
                        role=UserRole.PARENT,
                        first_name=first_name or "",
                        last_name=last_name or "",
                        is_active=True,
                    )
                else:
                    par = User.objects.create_user(
                        username=p_key,
                        email=None,
                        password=pwd,
                        role=UserRole.PARENT,
                        first_name="",
                        last_name="",
                        is_active=True,
                    )

                parent_cache[p_key] = par
                created_parent_passwords[p_key] = pwd
                return par, pwd

            with transaction.atomic():
                for worksheet in workbook.worksheets:
                    headers: dict[str, int] = {}
                    header_row: int | None = None

                    # Find header row on this sheet
                    for row_idx, row in enumerate(
                        worksheet.iter_rows(min_row=1, max_row=10, values_only=True),
                        start=1,
                    ):
                        if not row:
                            continue
                        normalized_cells = [str(cell).strip().lower() for cell in row if cell]
                        if not normalized_cells:
                            continue
                        if any(
                            c in ["class_name", "first_name", "last_name", "класс"]
                            or "фи ученика" in c
                            or "фио ученика" in c
                            or "почта" in c
                            or "фио родителя" in c
                            or "фи родителя" in c
                            for c in normalized_cells
                        ):
                            for col_idx, cell_value in enumerate(row, start=1):
                                if cell_value:
                                    norm_key = _normalize_header_name(cell_value)
                                    headers[norm_key] = col_idx
                            header_row = row_idx
                            break

                    if not headers or header_row is None:
                        continue
                    if "class_name" not in headers:
                        continue

                    has_separate_names = "first_name" in headers and "last_name" in headers
                    has_full_name = "student_full_name" in headers
                    if not has_separate_names and not has_full_name:
                        continue

                    # Process data rows on this sheet
                    for row_idx, row in enumerate(
                        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                        start=header_row + 1,
                    ):
                        if not any(cell for cell in row):
                            continue

                        # Extract data
                        class_name = (
                            str(row[headers["class_name"] - 1]).strip()
                            if headers["class_name"] <= len(row)
                            and row[headers["class_name"] - 1]
                            else None
                        )

                        # Student name: support both split and full-name formats
                        if "student_full_name" in headers and headers["student_full_name"] <= len(
                            row
                        ):
                            full_name_raw = row[headers["student_full_name"] - 1]
                            first_name = None
                            last_name = None
                            if full_name_raw:
                                last_name, first_name = split_full_name(full_name_raw)
                        else:
                            first_name = (
                                str(row[headers["first_name"] - 1]).strip()
                                if headers.get("first_name", 0) <= len(row)
                                and row[headers["first_name"] - 1]
                                else None
                            )
                            last_name = (
                                str(row[headers["last_name"] - 1]).strip()
                                if headers.get("last_name", 0) <= len(row)
                                and row[headers["last_name"] - 1]
                                else None
                            )

                        email = (
                            str(row[headers["email"] - 1]).strip()
                            if "email" in headers
                            and headers["email"] <= len(row)
                            and row[headers["email"] - 1]
                            else None
                        )
                        phone_number = (
                            str(row[headers["phone_number"] - 1]).strip()
                            if "phone_number" in headers
                            and headers["phone_number"] <= len(row)
                            and row[headers["phone_number"] - 1]
                            else None
                        )

                        # Parent: username (old format) or full name (new format)
                        parent_full_first = None
                        parent_full_last = None
                        if "parent_username" in headers:
                            row_parent = row_parent_username(
                                row, len(row), headers, default_parent_username
                            )
                        elif (
                            "parent_full_name" in headers
                            and headers["parent_full_name"] <= len(row)
                        ):
                            p_full = row[headers["parent_full_name"] - 1]
                            if p_full:
                                parent_full_last, parent_full_first = split_full_name(p_full)
                                row_parent = (
                                    f"{parent_full_last} {parent_full_first}".strip()
                                    if parent_full_last
                                    else None
                                )
                            else:
                                row_parent = None
                        else:
                            row_parent = None

                        # Validate required fields
                        if not class_name or not last_name:
                            # Строки без ФИ ученика пропускаем молча
                            continue

                        # Parse class_name to extract grade and letter
                        grade, letter = self._parse_class_name(class_name)
                        if not grade or not letter:
                            results["errors"].append(
                                {
                                    "row": row_idx,
                                    "error": f'Invalid class_name format: {class_name}. Expected format: "1A", "2Б", etc.',
                                }
                            )
                            continue

                        # Get or create classroom
                        classroom, created = Classroom.objects.get_or_create(
                            school=school,
                            grade=grade,
                            letter=letter,
                            defaults={"language": "ru"},
                        )

                        if class_name not in results["created_classrooms"]:
                            results["created_classrooms"][class_name] = 0

                        # Determine if student already exists:
                        # 1) by email + same school + student role
                        # 2) by first/last name + school + student role
                        existing_user = None
                        if email:
                            existing_user = User.objects.filter(
                                email=email,
                                school=school,
                                role=UserRole.STUDENT,
                            ).first()
                        if not existing_user:
                            existing_user = User.objects.filter(
                                first_name=first_name,
                                last_name=last_name,
                                school=school,
                                role=UserRole.STUDENT,
                            ).first()

                        # Each row in its own savepoint so one failure doesn't abort the whole tx
                        try:
                            with transaction.atomic():
                                if existing_user:
                                    # Attach to classroom if not already there
                                    already_in = ClassroomUser.objects.filter(
                                        classroom=classroom, user=existing_user
                                    ).exists()
                                    if not already_in:
                                        ClassroomUser.objects.create(
                                            classroom=classroom, user=existing_user
                                        )
                                        results["created_students"][class_name] = (
                                            results["created_students"].get(class_name, 0) + 1
                                        )

                                    # Always (re-)link parent even for existing students
                                    if row_parent:
                                        par_user, par_signal = get_or_create_parent(
                                            row_parent,
                                            first_name=parent_full_first,
                                            last_name=parent_full_last,
                                        )
                                        if par_user is None and par_signal == "not_parent":
                                            results["errors"].append(
                                                {
                                                    "row": row_idx,
                                                    "error": f'User "{row_parent}" exists but is not a parent',
                                                }
                                            )
                                        elif par_user and existing_user not in par_user.children.all():
                                            par_user.children.add(existing_user)
                                    continue

                                # New student
                                username = self._generate_username(first_name, last_name, school)
                                user = User.objects.create_user(
                                    username=username,
                                    email=email or None,
                                    password=default_password,
                                    role=UserRole.STUDENT,
                                    first_name=first_name,
                                    last_name=last_name,
                                    phone_number=phone_number if phone_number else None,
                                    school=school,
                                    is_active=True,
                                )

                                ClassroomUser.objects.create(
                                    classroom=classroom, user=user
                                )

                                parent_actual_username = None
                                if row_parent:
                                    par_user, par_signal = get_or_create_parent(
                                        row_parent,
                                        first_name=parent_full_first,
                                        last_name=parent_full_last,
                                    )
                                    if par_user is None and par_signal == "not_parent":
                                        results["errors"].append(
                                            {
                                                "row": row_idx,
                                                "error": f'User "{row_parent}" exists but is not a parent',
                                            }
                                        )
                                    elif par_user:
                                        par_user.children.add(user)
                                        parent_actual_username = par_user.username

                                results["new_students_credentials"].append(
                                    {
                                        "first_name": first_name,
                                        "last_name": last_name,
                                        "class_name": class_name,
                                        "username": username,
                                        "email": email,
                                        "password": default_password,
                                        "parent_username": parent_actual_username,
                                    }
                                )

                                results["created_students"][class_name] = (
                                    results["created_students"].get(class_name, 0) + 1
                                )

                        except Exception as e:
                            results["errors"].append(
                                {
                                    "row": row_idx,
                                    "error": f"Failed to process row: {str(e)}",
                                }
                            )
                            continue

            # After successful transaction, generate Excel with credentials if there are new students
            if results['new_students_credentials']:
                try:
                    import openpyxl
                    from openpyxl.utils import get_column_letter

                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Students credentials"

                    headers_row = [
                        "first_name",
                        "last_name",
                        "class_name",
                        "username",
                        "email",
                        "password",
                        "parent_username",
                    ]
                    ws.append(headers_row)

                    for cred in results['new_students_credentials']:
                        ws.append([
                            cred['first_name'],
                            cred['last_name'],
                            cred['class_name'],
                            cred['username'],
                            cred['email'],
                            cred['password'],
                            cred['parent_username'],
                        ])

                    for col_idx, column_title in enumerate(headers_row, start=1):
                        max_length = len(column_title)
                        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
                            cell_value = row[0]
                            if cell_value:
                                max_length = max(max_length, len(str(cell_value)))
                        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

                    media_root = Path(settings.MEDIA_ROOT)
                    target_dir = media_root / "import-credentials" / "students"
                    target_dir.mkdir(parents=True, exist_ok=True)

                    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
                    file_name = f"students_credentials_school_{school.id}_{timestamp}.xlsx"
                    credentials_file_path = target_dir / file_name
                    wb.save(credentials_file_path)
                except Exception as e:
                    results['errors'].append({
                        'row': None,
                        'error': f'Failed to generate credentials Excel file: {str(e)}',
                    })

            # Format response
            response_data = {
                'success': True,
                'message': 'Import completed',
                'summary': {
                    'total_classrooms': len(results['created_classrooms']),
                    'total_students': sum(results['created_students'].values()),
                    'errors_count': len(results['errors'])
                },
                'classrooms': [
                    {
                        'class_name': class_name,
                        'students_count': count
                    }
                    for class_name, count in results['created_students'].items()
                ],
                'default_password': results['default_password'],
                'errors': results['errors'][:50]  # Limit to first 50 errors
            }
            if created_parent_passwords:
                response_data['created_parents'] = [
                    {'username': uname, 'password': pwd}
                    for uname, pwd in created_parent_passwords.items()
                ]
            if credentials_file_path is not None:
                relative_path = credentials_file_path.relative_to(settings.MEDIA_ROOT)
                response_data['credentials_file'] = {
                    'path': str(relative_path),
                    'url': f"{settings.MEDIA_URL}{relative_path.as_posix()}",
                }

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': f'Error processing file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _parse_class_name(self, class_name):
        """
        Parse class_name like "1A", "2Б", "11Г" into (grade, letter)
        Returns (grade, letter) or (None, None) if invalid
        """
        if not class_name:
            return None, None

        # Remove whitespace
        class_name = class_name.strip()

        # Try to match pattern: number(s) followed by letter(s)
        match = re.match(r'^(\d+)([A-ZА-ЯЁа-яё]+)$', class_name, re.IGNORECASE)
        if match:
            grade_str = match.group(1)
            letter = match.group(2).upper()

            try:
                grade = int(grade_str)
                if 0 <= grade <= 12:
                    # Take only first letter if multiple letters provided
                    letter = letter[0]
                    return grade, letter
            except ValueError:
                pass

        return None, None

    def _transliterate_cyrillic_to_latin(self, text):
        """
        Transliterate Cyrillic (Russian + Kazakh) text to Latin.
        Supports Kazakh-specific letters: Ә, Ғ, Қ, Ң, Ө, Ұ, Ү, Һ, І
        """
        if not text:
            return ''

        # Transliteration mapping: Cyrillic -> Latin
        translit_map = {
            # Russian letters
            'А': 'A', 'а': 'a',
            'Б': 'B', 'б': 'b',
            'В': 'V', 'в': 'v',
            'Г': 'G', 'г': 'g',
            'Д': 'D', 'д': 'd',
            'Е': 'E', 'е': 'e',
            'Ё': 'Yo', 'ё': 'yo',
            'Ж': 'Zh', 'ж': 'zh',
            'З': 'Z', 'з': 'z',
            'И': 'I', 'и': 'i',
            'Й': 'Y', 'й': 'y',
            'К': 'K', 'к': 'k',
            'Л': 'L', 'л': 'l',
            'М': 'M', 'м': 'm',
            'Н': 'N', 'н': 'n',
            'О': 'O', 'о': 'o',
            'П': 'P', 'п': 'p',
            'Р': 'R', 'р': 'r',
            'С': 'S', 'с': 's',
            'Т': 'T', 'т': 't',
            'У': 'U', 'у': 'u',
            'Ф': 'F', 'ф': 'f',
            'Х': 'Kh', 'х': 'kh',
            'Ц': 'Ts', 'ц': 'ts',
            'Ч': 'Ch', 'ч': 'ch',
            'Ш': 'Sh', 'ш': 'sh',
            'Щ': 'Shch', 'щ': 'shch',
            'Ъ': '', 'ъ': '',  # Hard sign - remove
            'Ы': 'Y', 'ы': 'y',
            'Ь': '', 'ь': '',  # Soft sign - remove
            'Э': 'E', 'э': 'e',
            'Ю': 'Yu', 'ю': 'yu',
            'Я': 'Ya', 'я': 'ya',

            # Kazakh-specific letters
            'Ә': 'A', 'ә': 'a',  # A with diaeresis
            'Ғ': 'Gh', 'ғ': 'gh',  # G with stroke
            'Қ': 'Q', 'қ': 'q',  # K with descender
            'Ң': 'Ng', 'ң': 'ng',  # N with descender
            'Ө': 'O', 'ө': 'o',  # O with diaeresis
            'Ұ': 'U', 'ұ': 'u',  # U with stroke
            'Ү': 'U', 'ү': 'u',  # U with diaeresis
            'Һ': 'H', 'һ': 'h',  # H with descender
            'І': 'I', 'і': 'i',  # I with diaeresis
        }

        result = []
        for char in text:
            if char in translit_map:
                result.append(translit_map[char])
            elif char.isalnum() or char in ['-', '_', '.']:
                # Keep Latin letters, numbers, and common separators
                result.append(char)
            else:
                # Replace other characters with empty string or underscore
                result.append('')

        return ''.join(result)

    def _generate_username(self, first_name, last_name, school):
        """
        Generate unique username from first_name and last_name
        Format: lastname.firstname.number
        Supports Cyrillic (Russian + Kazakh) transliteration to Latin
        """
        # Transliterate Cyrillic to Latin
        last_name_translit = self._transliterate_cyrillic_to_latin(
            last_name).lower() if last_name else 'student'
        first_name_translit = self._transliterate_cyrillic_to_latin(
            first_name).lower() if first_name else 'x'

        # Remove any remaining non-alphanumeric characters except dots and hyphens
        allowed_chars = string.ascii_lowercase + string.digits + '.-'
        last_name_slug = ''.join(
            c for c in last_name_translit if c in allowed_chars)[:15]
        first_name_slug = ''.join(
            c for c in first_name_translit if c in allowed_chars)[:1]

        # Ensure we have valid slugs
        if not last_name_slug:
            last_name_slug = 'student'
        if not first_name_slug:
            first_name_slug = 'x'

        base_username = f"{last_name_slug}.{first_name_slug}"
        username = base_username
        counter = 1

        # Ensure uniqueness
        while User.objects.filter(username=username).exists():
            username = f"{base_username}{counter:03d}"
            counter += 1

        return username


class ClassroomViewSet(viewsets.ModelViewSet):
    queryset = Classroom.objects.select_related(
        'school').prefetch_related('classroom_users__user').all()
    serializer_class = ClassroomSerializer
    # Superadmins see и управляют всеми классами,
    # schooladmin — только классами своей школы
    permission_classes = [IsSchoolAdminOrSuperAdmin]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['school', 'grade', 'language']
    search_fields = ['letter', 'school__name']
    ordering_fields = ['grade', 'letter', 'school__name']
    ordering = ['school__name', 'grade', 'letter']

    def get_serializer_class(self):
        # Use detailed serializer for retrieve action (get single classroom)
        if self.action == 'retrieve':
            return ClassroomDetailSerializer
        return ClassroomSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        # Ограничиваем schooladmin только своей школой
        if getattr(user, "role", None) == UserRole.SCHOOLADMIN and user.school_id:
            queryset = queryset.filter(school_id=user.school_id)
        return queryset

    @action(detail=True, methods=['post'], url_path='add-student')
    def add_student(self, request, pk=None):
        """Add a single student to a classroom"""
        classroom = self.get_object()
        student_id = request.data.get('student_id')

        if not student_id:
            return Response(
                {'error': 'student_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            student = User.objects.get(id=student_id, role='student')
        except User.DoesNotExist:
            return Response(
                {'error': 'Student not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check if student is already in a classroom
        existing_classroom = ClassroomUser.objects.filter(user=student).first()
        if existing_classroom:
            return Response(
                {'error': f'Student is already in classroom {existing_classroom.classroom}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Add student to classroom
        classroom_user = ClassroomUser.objects.create(
            classroom=classroom,
            user=student
        )

        return Response(
            {
                'message': 'Student added successfully',
                'classroom_user_id': classroom_user.id
            },
            status=status.HTTP_201_CREATED
        )

    @action(detail=True, methods=['post'], url_path='remove-student')
    def remove_student(self, request, pk=None):
        """Remove a single student from a classroom"""
        classroom = self.get_object()
        student_id = request.data.get('student_id')

        if not student_id:
            return Response(
                {'error': 'student_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Find the ClassroomUser entry
            classroom_user = ClassroomUser.objects.get(
                classroom=classroom,
                user_id=student_id
            )
            classroom_user.delete()

            return Response(
                {'message': 'Student removed successfully'},
                status=status.HTTP_200_OK
            )
        except ClassroomUser.DoesNotExist:
            return Response(
                {'error': 'Student is not in this classroom'},
                status=status.HTTP_404_NOT_FOUND
            )


class ClassroomUserViewSet(viewsets.ModelViewSet):
    queryset = ClassroomUser.objects.select_related('classroom', 'user').all()
    serializer_class = ClassroomUserSerializer
    # Управление составом классов: супер‑ и школьные админы
    permission_classes = [IsSchoolAdminOrSuperAdmin]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['classroom', 'user']
    search_fields = ['user__username', 'user__email', 'classroom__letter']
    ordering_fields = ['user__username',
                       'classroom__grade', 'classroom__letter']
    ordering = ['classroom__school__name', 'classroom__grade',
                'classroom__letter', 'user__username']

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if getattr(user, "role", None) == UserRole.SCHOOLADMIN and user.school_id:
            queryset = queryset.filter(classroom__school_id=user.school_id)
        return queryset

    @action(detail=False, methods=['post'], url_path='bulk-add')
    def bulk_add(self, request):
        """Bulk add users to a classroom"""
        serializer = BulkClassroomUserSerializer(data=request.data)
        if serializer.is_valid():
            classroom_users = serializer.save()
            response_serializer = ClassroomUserSerializer(
                classroom_users, many=True)
            return Response(response_serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['delete'], url_path='bulk-remove')
    def bulk_remove(self, request):
        """Bulk remove users from a classroom"""
        classroom_id = request.data.get('classroom_id')
        user_ids = request.data.get('user_ids', [])

        if not classroom_id:
            return Response({'error': 'classroom_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = ClassroomUser.objects.filter(
            classroom_id=classroom_id,
            user_id__in=user_ids
        ).delete()

        return Response({'deleted_count': deleted_count}, status=status.HTTP_200_OK)
